from typing import Dict, Any, List, Optional, Union
from datetime import datetime, date
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from agno.tools import Toolkit
import logging

logger = logging.getLogger(__name__)


class ExcelFinanceManager(Toolkit):
    """Tool for managing financial data in Excel spreadsheets."""
    
    def __init__(self, file_path: Union[str, Path]):
        super().__init__()
        self.file_path = Path(file_path)
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self.initialize_workbook()
    
    def initialize_workbook(self):
        """Initialize the Excel workbook with necessary sheets."""
        try:
            if self.file_path.exists():
                self.workbook = load_workbook(self.file_path)
            else:
                self.workbook = Workbook()
                self.setup_initial_sheets()
                self.save_workbook()
        except Exception as e:
            logger.error(f"Failed to initialize workbook: {e}")
            self.workbook = Workbook()
            self.setup_initial_sheets()
    
    def setup_initial_sheets(self):
        """Set up initial sheets and headers."""
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # Create Expenses sheet
        expenses_sheet = self.workbook.create_sheet("Expenses")
        expenses_headers = [
            "Date", "Amount", "Category", "Description", 
            "Payment Method", "Notes", "Budget Impact"
        ]
        for col, header in enumerate(expenses_headers, 1):
            cell = expenses_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Create Budget sheet
        budget_sheet = self.workbook.create_sheet("Budget")
        budget_headers = [
            "Category", "Monthly Budget", "Current Spent", 
            "Remaining", "Percentage Used", "Status"
        ]
        for col, header in enumerate(budget_headers, 1):
            cell = budget_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Create Monthly Summary sheet
        summary_sheet = self.workbook.create_sheet("Monthly Summary")
        summary_headers = [
            "Month", "Total Income", "Total Expenses", 
            "Net Savings", "Top Category", "Budget Adherence %"
        ]
        for col, header in enumerate(summary_headers, 1):
            cell = summary_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Create Goals sheet
        goals_sheet = self.workbook.create_sheet("Goals")
        goals_headers = [
            "Goal Name", "Target Amount", "Current Amount", 
            "Deadline", "Progress %", "Status"
        ]
        for col, header in enumerate(goals_headers, 1):
            cell = goals_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    def add_expense(
        self,
        amount: float,
        category: str,
        description: str,
        payment_method: str = "Unknown",
        notes: str = "",
        expense_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """Add a new expense to the spreadsheet."""
        try:
            if expense_date is None:
                expense_date = datetime.now()
            
            expenses_sheet = self.workbook["Expenses"]
            
            # Find the next empty row
            next_row = expenses_sheet.max_row + 1
            
            # Add the expense data
            expenses_sheet.cell(row=next_row, column=1, value=expense_date.strftime("%Y-%m-%d"))
            expenses_sheet.cell(row=next_row, column=2, value=amount)
            expenses_sheet.cell(row=next_row, column=3, value=category)
            expenses_sheet.cell(row=next_row, column=4, value=description)
            expenses_sheet.cell(row=next_row, column=5, value=payment_method)
            expenses_sheet.cell(row=next_row, column=6, value=notes)
            
            # Calculate budget impact
            budget_impact = self.calculate_budget_impact(category, amount)
            expenses_sheet.cell(row=next_row, column=7, value=budget_impact)
            
            # Update budget tracking
            self.update_budget_tracking(category, amount)
            
            self.save_workbook()
            
            return {
                "success": True,
                "message": f"Added expense: ${amount:.2f} for {description}",
                "expense_id": next_row - 1,
                "budget_impact": budget_impact
            }
            
        except Exception as e:
            logger.error(f"Failed to add expense: {e}")
            return {"success": False, "error": str(e)}
    
    def update_budget_tracking(self, category: str, amount: float):
        """Update budget tracking for a category."""
        try:
            budget_sheet = self.workbook["Budget"]
            
            # Find if category exists in budget
            category_row = None
            for row in range(2, budget_sheet.max_row + 1):
                if budget_sheet.cell(row=row, column=1).value == category:
                    category_row = row
                    break
            
            if category_row:
                # Update existing category
                current_spent = budget_sheet.cell(row=category_row, column=3).value or 0
                new_spent = current_spent + amount
                monthly_budget = budget_sheet.cell(row=category_row, column=2).value or 0
                
                budget_sheet.cell(row=category_row, column=3, value=new_spent)
                budget_sheet.cell(row=category_row, column=4, value=monthly_budget - new_spent)
                
                if monthly_budget > 0:
                    percentage = (new_spent / monthly_budget) * 100
                    budget_sheet.cell(row=category_row, column=5, value=f"{percentage:.1f}%")
                    
                    # Set status based on percentage
                    if percentage > 100:
                        status = "OVER BUDGET"
                    elif percentage > 80:
                        status = "WARNING"
                    else:
                        status = "OK"
                    budget_sheet.cell(row=category_row, column=6, value=status)
            else:
                # Add new category to budget tracking
                next_row = budget_sheet.max_row + 1
                budget_sheet.cell(row=next_row, column=1, value=category)
                budget_sheet.cell(row=next_row, column=2, value=0)  # No budget set
                budget_sheet.cell(row=next_row, column=3, value=amount)
                budget_sheet.cell(row=next_row, column=4, value=-amount)
                budget_sheet.cell(row=next_row, column=5, value="N/A")
                budget_sheet.cell(row=next_row, column=6, value="NO BUDGET")
                
        except Exception as e:
            logger.error(f"Failed to update budget tracking: {e}")
    
    def calculate_budget_impact(self, category: str, amount: float) -> str:
        """Calculate the impact of an expense on the budget."""
        try:
            budget_sheet = self.workbook["Budget"]
            
            for row in range(2, budget_sheet.max_row + 1):
                if budget_sheet.cell(row=row, column=1).value == category:
                    monthly_budget = budget_sheet.cell(row=row, column=2).value or 0
                    current_spent = budget_sheet.cell(row=row, column=3).value or 0
                    
                    if monthly_budget > 0:
                        new_percentage = ((current_spent + amount) / monthly_budget) * 100
                        if new_percentage > 100:
                            return f"⚠️ Will exceed budget by ${(current_spent + amount) - monthly_budget:.2f}"
                        elif new_percentage > 80:
                            return f"⚠️ Will use {new_percentage:.1f}% of budget"
                        else:
                            return f"✅ Within budget ({new_percentage:.1f}% used)"
                    else:
                        return "ℹ️ No budget set for this category"
            
            return "ℹ️ New category - consider setting a budget"
            
        except Exception as e:
            logger.error(f"Failed to calculate budget impact: {e}")
            return "Error calculating impact"
    
    def set_budget(self, category: str, amount: float) -> Dict[str, Any]:
        """Set or update budget for a category."""
        try:
            budget_sheet = self.workbook["Budget"]
            
            # Find if category exists
            category_row = None
            for row in range(2, budget_sheet.max_row + 1):
                if budget_sheet.cell(row=row, column=1).value == category:
                    category_row = row
                    break
            
            if category_row:
                # Update existing budget
                budget_sheet.cell(row=category_row, column=2, value=amount)
                current_spent = budget_sheet.cell(row=category_row, column=3).value or 0
                budget_sheet.cell(row=category_row, column=4, value=amount - current_spent)
                
                if amount > 0:
                    percentage = (current_spent / amount) * 100
                    budget_sheet.cell(row=category_row, column=5, value=f"{percentage:.1f}%")
                    
                    if percentage > 100:
                        status = "OVER BUDGET"
                    elif percentage > 80:
                        status = "WARNING"
                    else:
                        status = "OK"
                    budget_sheet.cell(row=category_row, column=6, value=status)
            else:
                # Add new budget category
                next_row = budget_sheet.max_row + 1
                budget_sheet.cell(row=next_row, column=1, value=category)
                budget_sheet.cell(row=next_row, column=2, value=amount)
                budget_sheet.cell(row=next_row, column=3, value=0)
                budget_sheet.cell(row=next_row, column=4, value=amount)
                budget_sheet.cell(row=next_row, column=5, value="0.0%")
                budget_sheet.cell(row=next_row, column=6, value="OK")
            
            self.save_workbook()
            
            return {
                "success": True,
                "message": f"Budget set for {category}: ${amount:.2f}"
            }
            
        except Exception as e:
            logger.error(f"Failed to set budget: {e}")
            return {"success": False, "error": str(e)}
    
    def get_spending_summary(self, month: Optional[int] = None, year: Optional[int] = None) -> Dict[str, Any]:
        """Get spending summary for a specific month/year."""
        try:
            if month is None:
                month = datetime.now().month
            if year is None:
                year = datetime.now().year
            
            expenses_df = pd.read_excel(self.file_path, sheet_name="Expenses")
            expenses_df['Date'] = pd.to_datetime(expenses_df['Date'])
            
            # Filter by month and year
            month_expenses = expenses_df[
                (expenses_df['Date'].dt.month == month) & 
                (expenses_df['Date'].dt.year == year)
            ]
            
            total_spent = month_expenses['Amount'].sum()
            category_breakdown = month_expenses.groupby('Category')['Amount'].sum().to_dict()
            transaction_count = len(month_expenses)
            avg_transaction = month_expenses['Amount'].mean() if transaction_count > 0 else 0
            top_category = max(category_breakdown, key=category_breakdown.get) if category_breakdown else "None"
            
            return {
                "success": True,
                "month": month,
                "year": year,
                "total_spent": total_spent,
                "transaction_count": transaction_count,
                "average_transaction": avg_transaction,
                "category_breakdown": category_breakdown,
                "top_category": top_category,
                "top_category_amount": category_breakdown.get(top_category, 0)
            }
            
        except Exception as e:
            logger.error(f"Failed to get spending summary: {e}")
            return {"success": False, "error": str(e)}
    
    def get_budget_status(self) -> Dict[str, Any]:
        """Get current budget status for all categories."""
        try:
            budget_df = pd.read_excel(self.file_path, sheet_name="Budget")
            
            budget_status = []
            for _, row in budget_df.iterrows():
                budget_status.append({
                    "category": row['Category'],
                    "budget": row['Monthly Budget'],
                    "spent": row['Current Spent'],
                    "remaining": row['Remaining'],
                    "percentage": row['Percentage Used'],
                    "status": row['Status']
                })
            
            return {
                "success": True,
                "budget_status": budget_status
            }
            
        except Exception as e:
            logger.error(f"Failed to get budget status: {e}")
            return {"success": False, "error": str(e)}
    
    def set_user_balance(self, balance: float) -> Dict[str, Any]:
        """Set user's current balance in the User Setup sheet."""
        try:
            # Create User Setup sheet if it doesn't exist
            if "User Setup" not in self.workbook.sheetnames:
                setup_sheet = self.workbook.create_sheet("User Setup")
                setup_sheet.cell(row=1, column=1, value="Setting")
                setup_sheet.cell(row=1, column=2, value="Value")
                setup_sheet.cell(row=1, column=3, value="Last Updated")
                # Style headers
                for col in range(1, 4):
                    cell = setup_sheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            else:
                setup_sheet = self.workbook["User Setup"]
            
            # Find or create balance row
            balance_row = None
            for row in range(2, setup_sheet.max_row + 2):
                if setup_sheet.cell(row=row, column=1).value == "Current Balance":
                    balance_row = row
                    break
            
            if balance_row is None:
                balance_row = setup_sheet.max_row + 1
                setup_sheet.cell(row=balance_row, column=1, value="Current Balance")
            
            setup_sheet.cell(row=balance_row, column=2, value=balance)
            setup_sheet.cell(row=balance_row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
            
            self.save_workbook()
            
            return {
                "success": True,
                "message": f"Balance set to ${balance:,.2f}"
            }
            
        except Exception as e:
            logger.error(f"Failed to set user balance: {e}")
            return {"success": False, "error": str(e)}

    def get_user_setup_status(self) -> Dict[str, Any]:
        """Check if user has completed setup (balance and budgets)."""
        try:
            has_balance = False
            has_budgets = False
            
            # Check for balance in User Setup sheet
            if "User Setup" in self.workbook.sheetnames:
                setup_sheet = self.workbook["User Setup"]
                for row in range(2, setup_sheet.max_row + 1):
                    if setup_sheet.cell(row=row, column=1).value == "Current Balance":
                        balance_value = setup_sheet.cell(row=row, column=2).value
                        if balance_value is not None and balance_value > 0:
                            has_balance = True
                        break
            
            # Check for budgets in Budget sheet
            if "Budget" in self.workbook.sheetnames:
                budget_sheet = self.workbook["Budget"]
                for row in range(2, budget_sheet.max_row + 1):
                    budget_amount = budget_sheet.cell(row=row, column=2).value
                    if budget_amount is not None and budget_amount > 0:
                        has_budgets = True
                        break
            
            return {
                "has_balance": has_balance,
                "has_budgets": has_budgets,
                "setup_complete": has_balance and has_budgets
            }
            
        except Exception as e:
            logger.error(f"Failed to get user setup status: {e}")
            return {
                "has_balance": False,
                "has_budgets": False,
                "setup_complete": False
            }

    def set_category_budget(self, category: str, amount: float) -> Dict[str, Any]:
        """Set budget for a specific category (alias for set_budget with better naming)."""
        return self.set_budget(category, amount)

    def get_budget_setup_progress(self) -> Dict[str, Any]:
        """Get progress through budget setup process."""
        try:
            # This is a simplified implementation
            # In production, you'd track which categories have been configured
            priority_categories = [
                "Food & Dining",
                "Transportation", 
                "Shopping",
                "Bills & Utilities",
                "Entertainment"
            ]
            
            # Find which categories already have budgets
            configured_categories = []
            if "Budget" in self.workbook.sheetnames:
                budget_sheet = self.workbook["Budget"]
                for row in range(2, budget_sheet.max_row + 1):
                    category = budget_sheet.cell(row=row, column=1).value
                    budget_amount = budget_sheet.cell(row=row, column=2).value
                    if category and budget_amount and budget_amount > 0:
                        configured_categories.append(category)
            
            # Find next category to configure
            next_category = None
            for cat in priority_categories:
                if cat not in configured_categories:
                    next_category = cat
                    break
            
            return {
                "configured_categories": configured_categories,
                "current_category": next_category or priority_categories[0],
                "total_priority_categories": len(priority_categories),
                "configured_count": len(configured_categories)
            }
            
        except Exception as e:
            logger.error(f"Failed to get budget setup progress: {e}")
            return {
                "configured_categories": [],
                "current_category": "Food & Dining",
                "total_priority_categories": 5,
                "configured_count": 0
            }

    def mark_user_setup_complete(self) -> Dict[str, Any]:
        """Mark that user has completed the initial setup."""
        try:
            if "User Setup" not in self.workbook.sheetnames:
                self.set_user_balance(0)  # This will create the sheet
            
            setup_sheet = self.workbook["User Setup"]
            
            # Find or create setup complete row
            setup_row = None
            for row in range(2, setup_sheet.max_row + 2):
                if setup_sheet.cell(row=row, column=1).value == "Setup Complete":
                    setup_row = row
                    break
            
            if setup_row is None:
                setup_row = setup_sheet.max_row + 1
                setup_sheet.cell(row=setup_row, column=1, value="Setup Complete")
            
            setup_sheet.cell(row=setup_row, column=2, value="Yes")
            setup_sheet.cell(row=setup_row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
            
            self.save_workbook()
            
            return {"success": True, "message": "User setup marked as complete"}
            
        except Exception as e:
            logger.error(f"Failed to mark setup complete: {e}")
            return {"success": False, "error": str(e)}

    def get_user_budgets(self) -> Dict[str, Any]:
        """Get all user-configured budgets."""
        try:
            budgets = []
            
            if "Budget" in self.workbook.sheetnames:
                budget_sheet = self.workbook["Budget"]
                for row in range(2, budget_sheet.max_row + 1):
                    category = budget_sheet.cell(row=row, column=1).value
                    amount = budget_sheet.cell(row=row, column=2).value
                    
                    if category and amount and amount > 0:
                        budgets.append({
                            "category": category,
                            "amount": amount
                        })
            
            return {
                "success": True,
                "budgets": budgets
            }
            
        except Exception as e:
            logger.error(f"Failed to get user budgets: {e}")
            return {"success": False, "error": str(e)}

    def save_workbook(self):
        """Save the workbook to disk."""
        try:
            self.workbook.save(self.file_path)
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            raise